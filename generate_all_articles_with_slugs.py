#!/usr/bin/env python3
"""
Generate complete list of ALL articles with slugs.
This will be the final comprehensive check before handover.
"""

import os
import re
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl import Workbook

def extract_frontmatter(file_path):
    """Extract frontmatter from markdown file."""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        metadata = {}
        
        # Check for frontmatter (YAML between --- markers)
        frontmatter_match = re.match(r'^---\s*\n(.*?)\n---\s*\n(.*)$', content, re.DOTALL)
        if frontmatter_match:
            frontmatter_text = frontmatter_match.group(1)
            content = frontmatter_match.group(2)
            
            # Parse YAML-like frontmatter
            for line in frontmatter_text.split('\n'):
                line = line.strip()
                if ':' in line and not line.startswith('-'):
                    key, value = line.split(':', 1)
                    key = key.strip()
                    value = value.strip().strip('"').strip("'")
                    metadata[key] = value
        
        # If no title in frontmatter, try to extract from first heading
        if 'title' not in metadata:
            title_match = re.search(r'^#\s+(.+)$', content, re.MULTILINE)
            if title_match:
                metadata['title'] = title_match.group(1).strip()
            else:
                metadata['title'] = Path(file_path).stem.replace('-', ' ').replace('_', ' ').title()
        
        # Extract description if available
        if 'description' not in metadata:
            # Try to get first paragraph
            para_match = re.search(r'^([^\n]+)$', content.strip(), re.MULTILINE)
            if para_match:
                desc = para_match.group(1).strip()
                if len(desc) > 200:
                    desc = desc[:197] + '...'
                metadata['description'] = desc
        
        return metadata, content
    except Exception as e:
        return {'title': Path(file_path).stem.replace('-', ' ').replace('_', ' ').title()}, ''

def generate_slug_from_path(file_path, base_dir):
    """Generate slug from file path if not in frontmatter."""
    rel_path = os.path.relpath(file_path, base_dir)
    rel_path = rel_path.replace('.md', '').replace('.mdx', '')
    slug = rel_path.replace('\\', '/')
    
    if slug.endswith('/index') or slug.endswith('/README'):
        slug = slug.rsplit('/', 1)[0]
    
    return slug

def generate_url(slug):
    """Generate URL from slug."""
    return f"https://gcxone.info/docs/{slug}"

def get_category_from_path(file_path, base_dir):
    """Get category name from file path."""
    rel_path = os.path.relpath(file_path, base_dir)
    path_parts = rel_path.replace('\\', '/').split('/')
    
    if len(path_parts) == 1:
        return 'ROOT CATEGORY'
    elif len(path_parts) > 1:
        category = path_parts[0].replace('-', ' ').upper()
        return category
    return 'UNCATEGORIZED'

def scan_all_articles_with_slugs(docs_dir):
    """Scan all markdown files and extract articles with slugs."""
    articles = []
    
    for root, dirs, files in os.walk(docs_dir):
        # Skip certain directories
        if any(skip in root for skip in ['node_modules', '.git', '__pycache__', '.next']):
            continue
            
        for file in files:
            if file.endswith(('.md', '.mdx')):
                file_path = os.path.join(root, file)
                
                # Skip certain files
                filename_lower = file.lower()
                if filename_lower in ['readme.md', 'readme.mdx']:
                    continue
                
                metadata, content = extract_frontmatter(file_path)
                
                # Get slug - prefer from frontmatter, otherwise generate from path
                slug = metadata.get('slug')
                if not slug:
                    slug = generate_slug_from_path(file_path, docs_dir)
                
                # Skip if no meaningful slug
                if not slug or slug == '/':
                    continue
                
                title = metadata.get('title') or metadata.get('sidebar_label') or Path(file).stem
                description = metadata.get('description', '')
                category = get_category_from_path(file_path, docs_dir)
                url = generate_url(slug)
                
                article_info = {
                    'category': category,
                    'title': title,
                    'url': url,
                    'description': description,
                    'slug': slug,
                    'file': file,
                    'path': file_path
                }
                
                articles.append(article_info)
    
    return articles

def update_excel_file(articles, excel_path):
    """Update Excel file with all articles."""
    # Sort articles by category, then by title
    articles_sorted = sorted(articles, key=lambda x: (x['category'], x['title']))
    
    # Create or load workbook
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        # Clear existing data (keep header)
        ws.delete_rows(2, ws.max_row)
    else:
        wb = Workbook()
        ws = wb.active
        # Add header
        ws.append(['ROOT CATEGORY', 'Article Title', 'URL', 'Description'])
    
    # Add all articles
    for article in articles_sorted:
        ws.append([
            article['category'],
            article['title'],
            article['url'],
            article['description']
        ])
    
    # Save workbook
    wb.save(excel_path)
    print(f"Excel file updated: {excel_path}")
    print(f"Total articles: {len(articles_sorted)}")

def generate_markdown_summary(articles, output_file):
    """Generate markdown summary file."""
    articles_sorted = sorted(articles, key=lambda x: (x['category'], x['title']))
    
    # Group by category
    by_category = defaultdict(list)
    for article in articles_sorted:
        by_category[article['category']].append(article)
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("# Complete Article List - All Articles with Slugs\n\n")
        f.write("**Purpose**: Final comprehensive check before handover - ALL articles with slugs.\n\n")
        f.write("---\n\n")
        
        f.write(f"## Summary\n\n")
        f.write(f"**Total Articles**: {len(articles_sorted)}\n\n")
        f.write(f"**Categories**: {len(by_category)}\n\n")
        
        # Category breakdown
        f.write("### Category Breakdown\n\n")
        f.write("| Category | Count |\n")
        f.write("|----------|-------|\n")
        for category in sorted(by_category.keys()):
            count = len(by_category[category])
            f.write(f"| {category} | {count} |\n")
        
        f.write("\n---\n\n")
        
        # List by category
        for category in sorted(by_category.keys()):
            category_articles = by_category[category]
            f.write(f"## {category}\n\n")
            f.write(f"**Count**: {len(category_articles)}\n\n")
            f.write("| # | Title | URL | Slug |\n")
            f.write("|---|-------|-----|------|\n")
            
            for idx, article in enumerate(category_articles, 1):
                title = article['title'].replace('|', '\\|')
                url = article['url']
                slug = article['slug']
                f.write(f"| {idx} | {title} | [{url}]({url}) | `{slug}` |\n")
            
            f.write("\n---\n\n")

def main():
    """Main function."""
    base_dir = Path(__file__).parent
    docs_dir = base_dir / 'classic' / 'docs'
    excel_path = base_dir / 'Untitled Spreadsheet (2).xlsx'
    summary_file = base_dir / 'ALL_ARTICLES_WITH_SLUGS.md'
    
    if not docs_dir.exists():
        print(f"Error: Docs directory not found at {docs_dir}")
        return
    
    print(f"Scanning ALL documentation files: {docs_dir}")
    articles = scan_all_articles_with_slugs(str(docs_dir))
    
    print(f"Found {len(articles)} articles with slugs")
    
    # Update Excel file
    print(f"\nUpdating Excel file: {excel_path}")
    update_excel_file(articles, excel_path)
    
    # Generate markdown summary
    print(f"\nGenerating markdown summary: {summary_file}")
    generate_markdown_summary(articles, summary_file)
    
    print("\nComplete! All articles with slugs have been catalogued.")

if __name__ == '__main__':
    main()

